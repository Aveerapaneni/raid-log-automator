"""US-11 integration tests: prove raid_tool.py runs identically against an
xlsx-backed store, including cross-process persistence -- the actual
Section 14.7 DoD requirement, mirroring test_persistence_integration.py's
approach for the SQLite backend.

Each test shells out to raid_tool.py via subprocess.run() against a
working copy of the real RAID-log-template.xlsx (never the committed
template itself).
"""

import shutil
import subprocess
import sys
from pathlib import Path

REPO_DIR = Path(__file__).parent
RAID_TOOL = str(REPO_DIR / "raid_tool.py")
TEMPLATE = str(REPO_DIR / "RAID-log-template.xlsx")


def run(*args):
    result = subprocess.run(
        [sys.executable, RAID_TOOL, *args],
        capture_output=True,
        text=True,
        timeout=30,
    )
    return result.stdout + result.stderr, result.returncode


def xlsx_path(tmp_path, name="working_copy.xlsx"):
    return str(tmp_path / name)


def set_start_date(xlsx_file, entry_id, date_str):
    import openpyxl

    import raid_xlsx

    wb = openpyxl.load_workbook(xlsx_file)
    ws = wb[raid_xlsx.SHEET_NAME]
    for row in range(raid_xlsx.DATA_START_ROW, ws.max_row + 1):
        if ws.cell(row=row, column=raid_xlsx.COL_ID).value == entry_id:
            ws.cell(row=row, column=raid_xlsx.COL_START_DATE, value=date_str)
            break
    wb.save(xlsx_file)


# ---------------------------------------------------------------------------
# Every story runs against xlsx without error
# ---------------------------------------------------------------------------

def test_score_runs_against_xlsx(tmp_path):
    db = xlsx_path(tmp_path)
    output, code = run("score", "--data", TEMPLATE, "--db", db)
    assert code == 0
    assert "R-001" in output
    assert "I-003" in output  # blank-Probability Issue, unscored but valid


def test_status_runs_against_xlsx_with_no_auto_promotion(tmp_path):
    db = xlsx_path(tmp_path)
    output, code = run("status", "--data", TEMPLATE, "--db", db, "--today", "2026-08-15")
    assert code == 0
    assert "Status auto-promoted to In Progress: 0" in output


def test_materialize_runs_against_xlsx(tmp_path):
    db = xlsx_path(tmp_path)
    output, code = run("materialize", "--data", TEMPLATE, "--db", db)
    assert code == 0
    assert "Materialized-Risk conversions this run: 0" in output  # no materialized sample rows


def test_digest_runs_against_xlsx(tmp_path):
    db = xlsx_path(tmp_path)
    output, code = run("digest", "--data", TEMPLATE, "--db", db, "--count", "3")
    assert code == 0
    assert "RAID Log Digest" in output


def test_retain_close_and_remove_run_against_xlsx(tmp_path):
    db = xlsx_path(tmp_path)
    close_output, close_code = run("retain", "--data", TEMPLATE, "--db", db, "--close", "R-001")
    assert close_code == 0
    assert "R-001: Status set to Closed." in close_output

    remove_output, remove_code = run("retain", "--data", TEMPLATE, "--db", db, "--remove", "R-001")
    assert remove_code == 1
    assert "Refused" in remove_output


def test_sprint_ready_is_empty_without_start_date_but_populated_with_one(tmp_path):
    db = xlsx_path(tmp_path)
    run("score", "--data", TEMPLATE, "--db", db)  # seeds the working copy

    empty_output, _ = run("sprint-ready", "--data", TEMPLATE, "--db", db, "--today", "2026-08-15")
    assert "Sprint Ready pile is empty" in empty_output

    set_start_date(db, "R-001", "2026-08-01")
    populated_output, _ = run("sprint-ready", "--data", TEMPLATE, "--db", db, "--today", "2026-08-15")
    assert "R-001" in populated_output
    assert "Total Sprint Ready: 1" in populated_output


def test_reset_db_runs_against_xlsx(tmp_path):
    db = xlsx_path(tmp_path)
    run("retain", "--data", TEMPLATE, "--db", db, "--close", "R-001")
    output, code = run("reset-db", "--data", TEMPLATE, "--db", db)
    assert code == 0
    assert "reset to the 4 entries" in output


# ---------------------------------------------------------------------------
# Cross-process idempotency (Section 14.7 DoD) and template safety
# ---------------------------------------------------------------------------

def test_escalation_is_idempotent_across_separate_processes_against_xlsx(tmp_path):
    db = xlsx_path(tmp_path)
    common = ["escalate", "--data", TEMPLATE, "--db", db, "--today", "2026-08-15", "--score-band", "Low", "--days-open", "0", "--log", str(tmp_path / "log.jsonl")]

    first_output, first_code = run(*common)
    assert first_code == 0
    # 3 of the 4 sample rows escalate at Low/0 -- I-003 is already
    # "Escalated" in the sample data, correctly excluded by the
    # idempotency guard (breaches_threshold skips CLOSED/ESCALATED).
    assert "Escalated this run: 3" in first_output

    second_output, second_code = run(*common)
    assert second_code == 0
    assert "Escalated this run: 0" in second_output


def test_state_persists_across_separate_subcommand_invocations_on_xlsx(tmp_path):
    db = xlsx_path(tmp_path)
    run("retain", "--data", TEMPLATE, "--db", db, "--close", "D-004")

    status_output, status_code = run("status", "--data", TEMPLATE, "--db", db, "--today", "2026-08-15")
    assert status_code == 0
    assert "D-004" in status_output
    assert "Days Open blank (Closed): 1 -> ['D-004']" in status_output


def test_committed_template_is_never_modified_by_a_full_run(tmp_path):
    import hashlib

    before = hashlib.sha256(open(TEMPLATE, "rb").read()).hexdigest()

    db = xlsx_path(tmp_path)
    run("report", "--data", TEMPLATE, "--db", db, "--today", "2026-08-15")
    run("escalate", "--data", TEMPLATE, "--db", db, "--today", "2026-08-15", "--score-band", "Low", "--days-open", "0", "--log", str(tmp_path / "log.jsonl"))
    run("retain", "--data", TEMPLATE, "--db", db, "--close", "R-001")
    run("reset-db", "--data", TEMPLATE, "--db", db)

    after = hashlib.sha256(open(TEMPLATE, "rb").read()).hexdigest()
    assert before == after
