"""Unit tests for US-11's xlsx storage backend: raid_xlsx.py.

Uses the real RAID-log-template.xlsx from the repo as the seed (not a
synthetic fixture) -- the whole point of this backend is to work against
that specific file's actual layout, so testing against a stand-in would
miss real structural mismatches (exactly what caught the Start Date and
Status-vocabulary issues while building this).
"""

import openpyxl
import pytest

import raid_xlsx

TEMPLATE = "RAID-log-template.xlsx"


@pytest.fixture
def working_copy(tmp_path):
    path = str(tmp_path / "working_copy.xlsx")
    raid_xlsx.ensure_db(path, TEMPLATE)
    return path


# ---------------------------------------------------------------------------
# ensure_db
# ---------------------------------------------------------------------------

def test_ensure_db_creates_working_copy_from_template(tmp_path):
    path = str(tmp_path / "copy.xlsx")
    raid_xlsx.ensure_db(path, TEMPLATE)
    entries = raid_xlsx.load_entries(path)
    assert [e["id"] for e in entries] == ["R-001", "A-002", "I-003", "D-004"]


def test_ensure_db_adds_the_new_columns(working_copy):
    wb = openpyxl.load_workbook(working_copy)
    ws = wb[raid_xlsx.SHEET_NAME]
    assert ws.cell(row=raid_xlsx.HEADER_ROW, column=raid_xlsx.COL_MATERIALIZED).value == "Materialized"
    assert ws.cell(row=raid_xlsx.HEADER_ROW, column=raid_xlsx.COL_DEPENDENCY_LINKS).value == "Dependency Links"
    assert ws.cell(row=raid_xlsx.HEADER_ROW, column=raid_xlsx.COL_BLOCKED_BY).value == "Blocked By"
    assert ws.cell(row=raid_xlsx.HEADER_ROW, column=raid_xlsx.COL_START_DATE).value == "Start Date"


def test_ensure_db_is_a_no_op_if_working_copy_already_exists(working_copy):
    raid_xlsx.update_fields(working_copy, "R-001", status="Escalated")
    raid_xlsx.ensure_db(working_copy, TEMPLATE)  # must not re-copy over it
    [r001] = [e for e in raid_xlsx.load_entries(working_copy) if e["id"] == "R-001"]
    assert r001["status"] == "Escalated"


def test_ensure_db_never_modifies_the_committed_template():
    import hashlib

    before = hashlib.sha256(open(TEMPLATE, "rb").read()).hexdigest()
    import tempfile
    with tempfile.TemporaryDirectory() as d:
        raid_xlsx.ensure_db(f"{d}/copy.xlsx", TEMPLATE)
        raid_xlsx.update_fields(f"{d}/copy.xlsx", "R-001", status="Closed")
    after = hashlib.sha256(open(TEMPLATE, "rb").read()).hexdigest()
    assert before == after


# ---------------------------------------------------------------------------
# load_entries — real template data
# ---------------------------------------------------------------------------

def test_load_entries_reads_the_four_sample_rows_correctly(working_copy):
    by_id = {e["id"]: e for e in raid_xlsx.load_entries(working_copy)}
    r001 = by_id["R-001"]
    assert r001["category"] == "Risk"
    assert r001["owner"] == "Asha Veerpaneni"
    assert r001["probability"] == 3
    assert r001["impact"] == 4
    assert r001["status"] == "Monitoring"
    assert r001["date_raised"] == "2026-08-04"


def test_load_entries_blank_probability_stays_none_not_zero(working_copy):
    # I-003 is an Issue with intentionally blank Probability in the template.
    by_id = {e["id"]: e for e in raid_xlsx.load_entries(working_copy)}
    assert by_id["I-003"]["probability"] is None


def test_load_entries_start_date_defaults_to_none_when_blank(working_copy):
    for e in raid_xlsx.load_entries(working_copy):
        assert e["start_date"] is None  # none of the sample rows have one filled in


def test_load_entries_reads_a_manually_filled_in_start_date(working_copy):
    wb = openpyxl.load_workbook(working_copy)
    ws = wb[raid_xlsx.SHEET_NAME]
    ws.cell(row=5, column=raid_xlsx.COL_START_DATE, value="2026-08-01")
    wb.save(working_copy)

    [r001] = [e for e in raid_xlsx.load_entries(working_copy) if e["id"] == "R-001"]
    assert r001["start_date"] == "2026-08-01"


def test_load_entries_blank_status_defaults_to_open(working_copy):
    wb = openpyxl.load_workbook(working_copy)
    ws = wb[raid_xlsx.SHEET_NAME]
    ws.cell(row=5, column=raid_xlsx.COL_STATUS).value = None  # .cell(value=None) is a no-op, not a clear
    wb.save(working_copy)

    [r001] = [e for e in raid_xlsx.load_entries(working_copy) if e["id"] == "R-001"]
    assert r001["status"] == "Open"


def test_load_entries_materialized_blank_reads_as_none(working_copy):
    for e in raid_xlsx.load_entries(working_copy):
        assert e["materialized"] is None  # brand new column, nothing filled in yet


def test_load_entries_dependency_links_and_blocked_by_parse_comma_separated_ids(working_copy):
    wb = openpyxl.load_workbook(working_copy)
    ws = wb[raid_xlsx.SHEET_NAME]
    ws.cell(row=5, column=raid_xlsx.COL_DEPENDENCY_LINKS, value="A-002, D-004")
    ws.cell(row=5, column=raid_xlsx.COL_BLOCKED_BY, value="I-003")
    wb.save(working_copy)

    [r001] = [e for e in raid_xlsx.load_entries(working_copy) if e["id"] == "R-001"]
    assert r001["dependency_links"] == ["A-002", "D-004"]
    assert r001["blocked_by"] == ["I-003"]


def test_load_entries_never_reads_priority_score_or_days_open_as_input(working_copy):
    entries = raid_xlsx.load_entries(working_copy)
    for e in entries:
        assert "priority_score" not in e
        assert "days_open" not in e


def test_load_entries_raises_on_duplicate_id(working_copy):
    wb = openpyxl.load_workbook(working_copy)
    ws = wb[raid_xlsx.SHEET_NAME]
    ws.cell(row=6, column=raid_xlsx.COL_ID, value="R-001")  # duplicate of row 5
    wb.save(working_copy)

    with pytest.raises(ValueError, match="Duplicate ID"):
        raid_xlsx.load_entries(working_copy)


def test_load_entries_skips_blank_rows(working_copy):
    # Rows 9-28 in the template are formatted but empty.
    entries = raid_xlsx.load_entries(working_copy)
    assert len(entries) == 4


# ---------------------------------------------------------------------------
# update_fields
# ---------------------------------------------------------------------------

def test_update_fields_writes_status(working_copy):
    raid_xlsx.update_fields(working_copy, "R-001", status="Escalated")
    [r001] = [e for e in raid_xlsx.load_entries(working_copy) if e["id"] == "R-001"]
    assert r001["status"] == "Escalated"


def test_update_fields_writes_category(working_copy):
    raid_xlsx.update_fields(working_copy, "R-001", category="Issue")
    [r001] = [e for e in raid_xlsx.load_entries(working_copy) if e["id"] == "R-001"]
    assert r001["category"] == "Issue"


def test_update_fields_writes_materialized(working_copy):
    raid_xlsx.update_fields(working_copy, "R-001", materialized=True)
    [r001] = [e for e in raid_xlsx.load_entries(working_copy) if e["id"] == "R-001"]
    assert r001["materialized"] is True


def test_update_fields_never_touches_priority_score_or_days_open_cells(working_copy):
    wb_before = openpyxl.load_workbook(working_copy)
    formula_before = wb_before[raid_xlsx.SHEET_NAME].cell(row=5, column=8).value  # Priority Score

    raid_xlsx.update_fields(working_copy, "R-001", status="Closed")

    wb_after = openpyxl.load_workbook(working_copy)
    formula_after = wb_after[raid_xlsx.SHEET_NAME].cell(row=5, column=8).value
    assert formula_after == formula_before  # still the live formula, untouched


def test_update_fields_only_touches_the_named_entry(working_copy):
    raid_xlsx.update_fields(working_copy, "R-001", status="Closed")
    by_id = {e["id"]: e for e in raid_xlsx.load_entries(working_copy)}
    assert by_id["A-002"]["status"] == "Open"  # untouched


def test_update_fields_is_a_no_op_with_no_fields(working_copy):
    entries_before = raid_xlsx.load_entries(working_copy)
    raid_xlsx.update_fields(working_copy, "R-001")  # no keyword fields given
    entries_after = raid_xlsx.load_entries(working_copy)
    assert entries_before == entries_after


def test_update_fields_on_missing_id_does_not_crash(working_copy):
    raid_xlsx.update_fields(working_copy, "NO-SUCH-ID", status="Closed")  # silently no-ops


# ---------------------------------------------------------------------------
# reset_db
# ---------------------------------------------------------------------------

def test_reset_db_discards_persisted_state(working_copy):
    raid_xlsx.update_fields(working_copy, "R-001", status="Escalated", category="Issue")
    raid_xlsx.reset_db(working_copy, TEMPLATE)
    [r001] = [e for e in raid_xlsx.load_entries(working_copy) if e["id"] == "R-001"]
    assert r001["status"] == "Monitoring"
    assert r001["category"] == "Risk"


def test_reset_db_when_working_copy_does_not_exist_yet(tmp_path):
    path = str(tmp_path / "brand_new.xlsx")
    raid_xlsx.reset_db(path, TEMPLATE)
    entries = raid_xlsx.load_entries(path)
    assert [e["id"] for e in entries] == ["R-001", "A-002", "I-003", "D-004"]
