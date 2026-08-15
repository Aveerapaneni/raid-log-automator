"""US-11: xlsx storage backend for the RAID Log Automator (PRD Section 14).

Implements the same four-function contract as raid_db.py --
ensure_db/load_entries/update_fields/reset_db -- backed by an openpyxl
workbook instead of SQLite, so every other module works unchanged
regardless of which backend is active (see raid_store.py).

The committed RAID-log-template.xlsx is never written to. Automation
operates on a working copy created from it on first use. The working
copy's seed rows are whatever's already in the template (R-001, A-002,
...) -- its own independent example dataset, separate from
raid_mock_data.json (Section 4.1, 14.3).

Start Date IS added here (a column not in the original template) so
Sprint Ready (US-8) can work, which requires one. This doesn't reopen
the door to US-3's Status auto-promotion the PM asked to avoid:
compute_status() only promotes when Status is literally the string
"Not Started", and this backend's Status vocabulary is Open/Monitoring/
Escalated/Closed (preserved as-is from the template's own Legend sheet,
not remapped to the JSON schema's five values) -- so the trigger
condition can never be true here regardless of whether Start Date is
populated. Nothing in this codebase ever writes "Not Started" or
"In Progress" itself anyway (only US-4's "Escalated" and US-7's
"Closed" are ever written back), so no remapping is needed.

Priority Score and Days Open are never read as input NOR written as
output here: the template already has live Excel formulas computing
both correctly, and this tool doesn't persist either on the SQLite path
either -- they're always computed fresh for a report, never stored.
"""

import os
import shutil

import openpyxl

SHEET_NAME = "RAID Log"
HEADER_ROW = 4
DATA_START_ROW = 5

# 1-based column indices, matching RAID-log-template.xlsx exactly.
COL_ID = 1
COL_CATEGORY = 2
COL_DESCRIPTION = 3
COL_DATE_RAISED = 4
COL_OWNER = 5
COL_PROBABILITY = 6
COL_IMPACT = 7
# COL 8 = Priority Score -- calculated, never read or written here.
COL_STATUS = 9
COL_MITIGATION_PLAN = 10
COL_TARGET_DATE = 11
# COL 12 = Days Open -- calculated, never read or written here.
COL_LAST_UPDATED = 13
# COL 14 = Notes -- a PM's free-text field, never touched here.

# New columns appended by ensure_db, at columns 15-18.
COL_MATERIALIZED = 15
COL_DEPENDENCY_LINKS = 16
COL_BLOCKED_BY = 17
COL_START_DATE = 18
NEW_COLUMNS = {
    COL_MATERIALIZED: "Materialized",
    COL_DEPENDENCY_LINKS: "Dependency Links",
    COL_BLOCKED_BY: "Blocked By",
    COL_START_DATE: "Start Date",
}

WRITABLE_COLUMNS = {"status": COL_STATUS, "category": COL_CATEGORY, "materialized": COL_MATERIALIZED}
DEFAULT_STATUS = "Open"


def ensure_db(file_path, template_path):
    """Creates the working copy from template_path if file_path doesn't
    exist yet. No-op if it already exists -- mirrors raid_db.ensure_db."""
    if os.path.exists(file_path):
        return
    shutil.copy(template_path, file_path)
    wb = openpyxl.load_workbook(file_path)
    ws = wb[SHEET_NAME]
    for col, header in NEW_COLUMNS.items():
        ws.cell(row=HEADER_ROW, column=col, value=header)
    wb.save(file_path)


def _date_to_str(value):
    if value is None or value == "":
        return None
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)


def _split_ids(value):
    if not value:
        return []
    return [v.strip() for v in str(value).split(",") if v.strip()]


def _row_to_entry(ws, row, entry_id):
    def cell(col):
        return ws.cell(row=row, column=col).value

    status = cell(COL_STATUS)
    materialized_raw = cell(COL_MATERIALIZED)

    return {
        "id": entry_id,
        "category": cell(COL_CATEGORY),
        "description": cell(COL_DESCRIPTION),
        "date_raised": _date_to_str(cell(COL_DATE_RAISED)),
        "owner": cell(COL_OWNER),
        "probability": cell(COL_PROBABILITY),
        "impact": cell(COL_IMPACT),
        "mitigation_plan": cell(COL_MITIGATION_PLAN),
        "start_date": _date_to_str(cell(COL_START_DATE)),
        "status": status if status not in (None, "") else DEFAULT_STATUS,
        "materialized": bool(materialized_raw) if materialized_raw not in (None, "") else None,
        "dependency_links": _split_ids(cell(COL_DEPENDENCY_LINKS)),
        "blocked_by": _split_ids(cell(COL_BLOCKED_BY)),
        "target_date": _date_to_str(cell(COL_TARGET_DATE)),
        "last_updated": _date_to_str(cell(COL_LAST_UPDATED)),
    }


def load_entries(file_path):
    """Returns all entries as plain dicts matching the same shape used
    everywhere else in the codebase, in sheet order. Raises on a
    duplicate ID (Section 14.6) -- xlsx has no primary-key constraint to
    catch this for free the way SQLite does."""
    wb = openpyxl.load_workbook(file_path)
    ws = wb[SHEET_NAME]

    entries = []
    seen_ids = set()
    for row in range(DATA_START_ROW, ws.max_row + 1):
        entry_id = ws.cell(row=row, column=COL_ID).value
        if entry_id in (None, ""):
            continue
        entry_id = str(entry_id).strip()
        if entry_id in seen_ids:
            raise ValueError(
                f"Duplicate ID {entry_id!r} in {file_path!r} (row {row}) -- "
                "xlsx has no primary-key constraint; fix the sheet before continuing."
            )
        seen_ids.add(entry_id)
        entries.append(_row_to_entry(ws, row, entry_id))
    return entries


def _find_row(ws, entry_id):
    for row in range(DATA_START_ROW, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=COL_ID).value
        if cell_value not in (None, "") and str(cell_value).strip() == entry_id:
            return row
    return None


def update_fields(file_path, entry_id, **fields):
    """Writes back only Status/Category/Materialized -- the identical
    three fields raid_db.update_fields() writes, and nothing else.
    No-op if fields is empty or entry_id isn't found."""
    if not fields:
        return
    wb = openpyxl.load_workbook(file_path)
    ws = wb[SHEET_NAME]
    row = _find_row(ws, entry_id)
    if row is None:
        return
    for field, value in fields.items():
        col = WRITABLE_COLUMNS.get(field)
        if col is None:
            continue
        ws.cell(row=row, column=col, value=value)
    wb.save(file_path)


def reset_db(file_path, template_path):
    """US-10 parity: discard the working copy and reseed fresh from
    template_path. Never modifies template_path itself."""
    if os.path.exists(file_path):
        os.remove(file_path)
    ensure_db(file_path, template_path)
